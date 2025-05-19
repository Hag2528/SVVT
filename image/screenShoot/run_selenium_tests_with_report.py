import subprocess
import os
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import glob
import unittest
import xmlrunner
from selenium_tests import ContractRenewalSystemTest
import argparse
import shutil
import webbrowser

def run_selenium_tests():
    """Run Selenium UI tests with XML report"""
    print("=" * 80)
    print("RUNNING SELENIUM UI TESTS")
    print("=" * 80)
    
    try:
        # Create screenshots directory if it doesn't exist
        screenshots_dir = "test_reports/screenshots"
        if not os.path.exists(screenshots_dir):
            os.makedirs(screenshots_dir)
        
        # Create reports directory if it doesn't exist
        if not os.path.exists('test_reports'):
            os.makedirs('test_reports')
        
        # Generate timestamp for report directory
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        report_dir = f"test_reports/selenium_test_report_{timestamp}"
        
        # Set environment variable to enable screenshots
        os.environ['TAKE_SCREENSHOTS'] = 'True'
        os.environ['SCREENSHOTS_DIR'] = screenshots_dir
        
        # Run tests with XML reporter
        suite = unittest.TestLoader().loadTestsFromTestCase(ContractRenewalSystemTest)
        runner = xmlrunner.XMLTestRunner(output=report_dir)
        result = runner.run(suite)
        
        # Parse XML results
        xml_results = []
        xml_files = glob.glob(f"{report_dir}/*.xml")
        
        if xml_files:
            latest_xml = max(xml_files, key=os.path.getctime)
            try:
                tree = ET.parse(latest_xml)
                root = tree.getroot()
                
                for testcase in root.findall('.//testcase'):
                    test_name = testcase.get('name')
                    test_class = testcase.get('classname')
                    test_time = float(testcase.get('time', 0))
                    
                    # Check if test failed
                    failure = testcase.find('failure')
                    error = testcase.find('error')
                    
                    status = "PASSED"
                    error_message = ""
                    
                    if failure is not None:
                        status = "FAILED"
                        error_message = failure.get('message', '')
                    elif error is not None:
                        status = "ERROR"
                        error_message = error.get('message', '')
                    
                    # Find screenshot for this test
                    screenshots = []
                    for screenshot_file in glob.glob(f"{screenshots_dir}/*.png"):
                        if test_name.replace("test_", "") in os.path.basename(screenshot_file):
                            screenshots.append(screenshot_file)
                    
                    xml_results.append({
                        "name": f"{test_class}.{test_name}",
                        "status": status,
                        "time": test_time,
                        "reason": error_message,
                        "screenshots": screenshots
                    })
            except Exception as e:
                print(f"Error parsing XML results: {e}")
        
        success = result.wasSuccessful()
        if success:
            print("\n✅ Selenium UI tests passed!")
        else:
            print("\n❌ Selenium UI tests failed.")
        
        return {
            "name": "Selenium UI Tests",
            "status": "PASSED" if success else "FAILED",
            "exit_code": 0 if success else 1,
            "detailed_results": xml_results,
            "screenshots_dir": screenshots_dir
        }
        
    except Exception as e:
        print(f"Error running Selenium tests: {e}")
        return {
            "name": "Selenium UI Tests",
            "status": "ERROR",
            "exit_code": 1,
            "detailed_results": []
        }

def generate_excel_report(results):
    """Generate Excel report with test results in tabular format"""
    print("\n" + "=" * 80)
    print("GENERATING EXCEL REPORT")
    print("=" * 80)
    
    # Create a new workbook
    wb = Workbook()
    
    # Create summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Test Summary"
    
    # Add headers
    summary_sheet['A1'] = "Test Type"
    summary_sheet['B1'] = "Status"
    summary_sheet['C1'] = "Total Tests"
    summary_sheet['D1'] = "Passed"
    summary_sheet['E1'] = "Failed"
    summary_sheet['F1'] = "Error"
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    row = 2
    summary_sheet[f'A{row}'] = results["name"]
    summary_sheet[f'B{row}'] = results["status"]
    
    # Calculate statistics
    if "detailed_results" in results and results["detailed_results"]:
        total_tests = len(results["detailed_results"])
        passed_tests = sum(1 for r in results["detailed_results"] if r["status"] == "PASSED")
        failed_tests = sum(1 for r in results["detailed_results"] if r["status"] == "FAILED")
        error_tests = sum(1 for r in results["detailed_results"] if r["status"] == "ERROR")
        
        summary_sheet[f'C{row}'] = total_tests
        summary_sheet[f'D{row}'] = passed_tests
        summary_sheet[f'E{row}'] = failed_tests
        summary_sheet[f'F{row}'] = error_tests
    else:
        summary_sheet[f'C{row}'] = "N/A"
        summary_sheet[f'D{row}'] = "N/A"
        summary_sheet[f'E{row}'] = "N/A"
        summary_sheet[f'F{row}'] = "N/A"
    
    # Style based on status
    if results["status"] == "PASSED":
        summary_sheet[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif results["status"] == "FAILED":
        summary_sheet[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    else:
        summary_sheet[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Auto-size columns
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create detailed results sheet
    if "detailed_results" in results and results["detailed_results"]:
        detailed_sheet = wb.create_sheet(title="Test_Details")
        
        # Add headers
        detailed_sheet['A1'] = "Test Name"
        detailed_sheet['B1'] = "Status"
        detailed_sheet['C1'] = "Time (s)"
        detailed_sheet['D1'] = "Reason/Error"
        detailed_sheet['E1'] = "Screenshots"
        
        # Style headers
        for cell in detailed_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        row = 2
        for test_result in results["detailed_results"]:
            detailed_sheet[f'A{row}'] = test_result["name"]
            detailed_sheet[f'B{row}'] = test_result["status"]
            detailed_sheet[f'C{row}'] = test_result["time"]
            detailed_sheet[f'D{row}'] = test_result.get("reason", "")
            
            # Add screenshot info
            if "screenshots" in test_result and test_result["screenshots"]:
                screenshot_info = ", ".join([os.path.basename(s) for s in test_result["screenshots"]])
                detailed_sheet[f'E{row}'] = screenshot_info
            else:
                detailed_sheet[f'E{row}'] = "No screenshots"
            
            # Style based on status
            if test_result["status"] == "PASSED":
                detailed_sheet[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif test_result["status"] == "FAILED":
                detailed_sheet[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                detailed_sheet[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            row += 1
        
        # Auto-size columns
        for column in detailed_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50
            detailed_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate timestamp for report filename
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = f"test_reports/selenium_test_report_{timestamp}.xlsx"
    
    # Save the workbook
    wb.save(report_file)
    
    print(f"Excel report generated: {report_file}")
    
    # Automatically open the Excel file
    try:
        import os
        import platform
        
        if platform.system() == 'Windows':
            os.system(f'start excel.exe "{report_file}"')
        elif platform.system() == 'Darwin':  # macOS
            os.system(f'open "{report_file}"')
        else:  # Linux
            os.system(f'xdg-open "{report_file}"')
        
        print(f"Excel report opened automatically")
    except Exception as e:
        print(f"Could not open Excel report automatically: {e}")
        print(f"Please open the file manually: {report_file}")
    
    return report_file

def copy_to_downloads(report_file):
    """Copy the report file to the Downloads folder"""
    try:
        # Get the Downloads folder path
        if os.name == 'nt':  # Windows
            import winreg
            sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
            downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
                downloads_path = winreg.QueryValueEx(key, downloads_guid)[0]
        else:  # macOS and Linux
            downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
        
        # Ensure the Downloads folder exists
        if not os.path.exists(downloads_path):
            print(f"Downloads folder not found at {downloads_path}")
            return None
        
        # Copy the file to Downloads
        filename = os.path.basename(report_file)
        download_file = os.path.join(downloads_path, filename)
        shutil.copy2(report_file, download_file)
        
        print(f"Report copied to Downloads folder: {download_file}")
        return download_file
    
    except Exception as e:
        print(f"Error copying report to Downloads folder: {e}")
        return None

def main():
    """Run Selenium tests and generate Excel report"""
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Run Selenium tests and generate Excel report')
    parser.add_argument('--open', action='store_true', help='Open Excel report after generation')
    parser.add_argument('--download', action='store_true', help='Copy report to Downloads folder')
    args = parser.parse_args()
    
    # Run Selenium tests
    results = run_selenium_tests()
    
    # Generate Excel report
    report_file = generate_excel_report(results)
    
    # Copy to Downloads folder if requested
    if args.download:
        download_file = copy_to_downloads(report_file)
        if download_file:
            # Open the downloaded file instead if --open is also specified
            if args.open:
                try:
                    webbrowser.open('file://' + os.path.abspath(download_file))
                    print(f"Opened downloaded report: {download_file}")
                except Exception as e:
                    print(f"Could not open downloaded report: {e}")
    
    print("\n" + "=" * 80)
    print(f"TESTING COMPLETE - Report: {report_file}")
    print("=" * 80)
    
    return results["exit_code"]

if __name__ == "__main__":
    exit_code = main()
    exit(exit_code)



